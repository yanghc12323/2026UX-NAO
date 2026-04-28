# -*- coding: utf-8 -*-
"""实验 Web 控制台后端（标准库实现，零额外 Web 框架依赖）。

能力概览：
1) 主试录入被试信息 + 选择 2x2 条件；
2) 手动切换实验阶段（warmup/task_intro/formal_interview/closing_and_questionnaire）；
3) 接收 Python2 推送器实时数据（/asr, /gaze），并计算核心指标；
4) 前端轮询状态；
5) 会话结束后导出本地 Excel/CSV（优先 xlsx，缺少 openpyxl 时回退 csv）。
"""

from __future__ import annotations

import argparse
import json
import os
import threading
import time
import uuid
from dataclasses import dataclass, field
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from typing import Dict, List, Optional
from urllib.error import URLError, HTTPError
from urllib import request

from client.metrics import compute_fluency_metrics, compute_gaze_contact_ratio
from client.interview_policy import InterviewPolicy, ConditionPolicy
from client.llm_interview_provider import LLMQuestionProvider, LLMFeedbackProvider
from client.llm_provider import LLMClient, LLMConfig


STAGES = [
    "warmup",
    "task_intro",
    "formal_interview",
    "closing_and_questionnaire",
]

STAGE_ENTRY_SCRIPT = {
    "warmup": {
        "speak": "欢迎你来，我们先轻松聊一会儿，帮你放松一下。你最近在学校里有什么让你开心的小事吗？",
        "nod": True,
    },
    "task_intro": {
        "speak": "接下来是任务说明：你将先做1分钟自我介绍，然后回答4道行为面试题。",
        "nod": False,
    },
    "formal_interview": {
        "speak": "现在进入正式面试阶段。请先用约1分钟做自我介绍，然后我们开始行为问题。",
        "nod": True,
    },
    "closing_and_questionnaire": {
        "speak": "今天的模拟面试到这里，感谢你的参与。请继续完成后续问卷。",
        "nod": False,
    },
}

CONDITION_MATRIX = {
    "C1": {"persona_style": "encouraging", "backchanneling_type": "positive", "label": "鼓励型 × 积极反馈"},
    "C2": {"persona_style": "encouraging", "backchanneling_type": "negative", "label": "鼓励型 × 消极反馈"},
    "C3": {"persona_style": "pressure", "backchanneling_type": "positive", "label": "压力型 × 积极反馈"},
    "C4": {"persona_style": "pressure", "backchanneling_type": "negative", "label": "压力型 × 消极反馈"},
}

ASR_STALE_TIMEOUT_MS = 10000
GAZE_STALE_TIMEOUT_MS = 10000
ROBOT_HEALTH_CACHE_MS = 3000
STAGE_SYNC_GRACE_MS = 1500


def now_ms() -> int:
    return int(time.time() * 1000)


@dataclass
class ParticipantSession:
    session_id: str
    participant_id: str
    participant_name: str
    condition_id: str
    persona_style: str
    backchanneling_type: str
    started_at_ms: int
    ended_at_ms: int = 0
    current_stage: str = "warmup"
    stage_history: List[dict] = field(default_factory=list)


class DialogueOrchestrator(object):
    """Web 侧唯一对话编排器：输入 ASR+stage+condition，输出动作与状态推进。"""

    def __init__(self, send_robot_command, question_provider, feedback_provider):
        self.send_robot_command = send_robot_command
        self.question_provider = question_provider
        self.feedback_provider = feedback_provider

    def on_asr(self, stage: str, user_text: str, dialogue_state: dict) -> dict:
        if stage == "warmup":
            print(f"[DEBUG] DialogueOrchestrator.on_asr: stage=warmup, user_text={user_text[:50]}...")
            reply = self._warmup_reply_text(
                user_text,
                fallback="谢谢你分享，我们先轻松聊聊～最近在学校里有没有什么让你开心的小事？",
            )
            print(f"[DEBUG] LLM reply generated: {reply[:50]}...")
            resp = self.send_robot_command("speak", {"text": reply})
            print(f"[DEBUG] Robot command response: status={resp.get('status')}")
            return {
                "triggered": True,
                "actions": [{"command": "speak", "text": reply, "ok": str(resp.get("status", "")).lower() == "ok", "tag": "warmup_reply"}],
                "next_state": dialogue_state,
            }

        if stage != "formal_interview":
            return {"triggered": False, "reason": "stage_not_enabled", "next_state": dialogue_state}

        step = str((dialogue_state or {}).get("formal_step", "formal_self_intro_wait"))
        actions = []
        questions = list((dialogue_state or {}).get("star_questions", []))
        if not questions:
            questions = self._main_questions()[:4]

        if step == "formal_self_intro_wait":
            fb = self._feedback_text(user_text, fallback="收到。建议你补充更具体的行动和量化结果。")
            r1 = self.send_robot_command("speak", {"text": fb})
            actions.append({"command": "speak", "text": fb, "ok": str(r1.get("status", "")).lower() == "ok", "tag": "self_intro_feedback"})
            q1 = questions[0] if questions else "请用 STAR 结构介绍一个你主导推进并最终落地的项目经历。"
            r2 = self.send_robot_command("speak", {"text": q1})
            actions.append({"command": "speak", "text": q1, "ok": str(r2.get("status", "")).lower() == "ok", "tag": "star_q1"})
            next_state = dict(dialogue_state or {})
            next_state.update({"formal_step": "formal_star_q1_wait", "star_questions": questions, "star_index": 0})
            return {"triggered": True, "actions": actions, "next_state": next_state}

        step_map = {
            "formal_star_q1_wait": 0,
            "formal_star_q2_wait": 1,
            "formal_star_q3_wait": 2,
            "formal_star_q4_wait": 3,
        }
        if step in step_map:
            idx = step_map[step]
            fb = self._feedback_text(user_text, fallback="收到。建议你补充更具体的行动和量化结果。")
            r1 = self.send_robot_command("speak", {"text": fb})
            actions.append({"command": "speak", "text": fb, "ok": str(r1.get("status", "")).lower() == "ok", "tag": "star_feedback"})
            next_idx = idx + 1
            next_state = dict(dialogue_state or {})
            next_state.update({"star_questions": questions, "star_index": next_idx})
            if next_idx < min(4, len(questions)):
                next_q = questions[next_idx]
                r2 = self.send_robot_command("speak", {"text": next_q})
                actions.append({"command": "speak", "text": next_q, "ok": str(r2.get("status", "")).lower() == "ok", "tag": "star_next"})
                next_state["formal_step"] = "formal_star_q%d_wait" % (next_idx + 1)
            else:
                done_text = "正式问答部分已完成。请等待主试进入结束阶段。"
                r3 = self.send_robot_command("speak", {"text": done_text})
                actions.append({"command": "speak", "text": done_text, "ok": str(r3.get("status", "")).lower() == "ok", "tag": "formal_done"})
                next_state["formal_step"] = "formal_done"
            return {"triggered": True, "actions": actions, "next_state": next_state}

        return {"triggered": False, "reason": "no_transition", "next_state": dialogue_state}

    def _main_questions(self):
        if self.question_provider is None:
            return []
        return self.question_provider.get_main_questions()

    def _feedback_text(self, answer_text: str, fallback: str) -> str:
        if self.feedback_provider is None:
            return fallback
        try:
            return self.feedback_provider.feedback_for_answer(answer_text)
        except Exception:
            return fallback

    def _warmup_reply_text(self, user_text: str, fallback: str) -> str:
        print(f"[DEBUG] _warmup_reply_text called, provider={self.feedback_provider}")
        if self.feedback_provider is None:
            print("[DEBUG] No feedback_provider, using fallback")
            return fallback
        try:
            # 若提供器支持 warmup 轻松聊天回复，则优先使用。
            if hasattr(self.feedback_provider, "warmup_reply"):
                print("[DEBUG] Calling feedback_provider.warmup_reply()")
                reply = self.feedback_provider.warmup_reply(user_text)
                print(f"[DEBUG] Got reply from LLM: {reply[:50]}...")
                return reply
            print("[DEBUG] No warmup_reply method, using feedback_for_answer")
            return self.feedback_provider.feedback_for_answer(user_text)
        except Exception as e:
            print(f"[DEBUG] Exception in _warmup_reply_text: {e}")
            return fallback


class ExperimentState(object):
    """线程安全的实验状态容器。"""

    def __init__(self, export_dir: str, robot_server_url: str):
        self._lock = threading.Lock()
        self.export_dir = export_dir
        self.robot_server_url = robot_server_url

        self.session: Optional[ParticipantSession] = None
        self.latest_gaze_by_stage: Dict[str, dict] = {}
        self.stage_gaze_offset_by_stage: Dict[str, float] = {}
        self.latest_raw_gaze_total_s: float = 0.0
        self.latest_metrics: Optional[dict] = None
        self.metric_events: List[dict] = []
        self.asr_events: List[dict] = []
        self.gaze_events: List[dict] = []
        self.last_asr_received_at_ms: int = 0
        self.last_gaze_received_at_ms: int = 0
        self.stage_changed_at_ms: int = 0
        self.stage_version: int = 0
        self._robot_health_cache: dict = {
            "checked_at_ms": 0,
            "ok": False,
            "message": "not_checked",
        }
        self.dialogue_state: Dict[str, object] = {
            "formal_step": "formal_self_intro_wait",
            "star_questions": [],
            "star_index": 0,
            "self_intro_prompt": "",
        }

        # LLM 组件：失败时自动回退，不中断实验流程。
        self._llm = LLMClient(LLMConfig())
        self._question_provider: Optional[LLMQuestionProvider] = None
        self._feedback_provider: Optional[LLMFeedbackProvider] = None
        self._llm_prewarm_done: bool = False
        self._llm_prewarm_status: dict = {"ok": False, "message": "not_started", "timestamp_ms": 0}

    def _build_policy_from_condition(self, condition_id: str) -> InterviewPolicy:
        cp = ConditionPolicy.from_condition_id(condition_id)
        return cp.to_interview_policy()

    def start_session(self, participant_id: str, participant_name: str, condition_id: str) -> dict:
        condition = CONDITION_MATRIX.get(condition_id)
        if not condition:
            raise ValueError("invalid_condition_id")
        if not participant_id.strip():
            raise ValueError("empty_participant_id")
        if not participant_name.strip():
            raise ValueError("empty_participant_name")

        sid = "S_%s" % uuid.uuid4().hex[:10].upper()
        created = ParticipantSession(
            session_id=sid,
            participant_id=participant_id.strip(),
            participant_name=participant_name.strip(),
            condition_id=condition_id,
            persona_style=condition["persona_style"],
            backchanneling_type=condition["backchanneling_type"],
            started_at_ms=now_ms(),
        )
        created.stage_history.append({"stage": created.current_stage, "timestamp_ms": now_ms(), "source": "session_start"})

        with self._lock:
            # 保留会话启动前最近一次 ASR/Gaze 心跳时间，避免“开始会话”瞬间把连接状态清零，
            # 导致前端立刻误报 ASR/GAZE 断开。
            preserved_last_asr_ms = self.last_asr_received_at_ms
            preserved_last_gaze_ms = self.last_gaze_received_at_ms

            self.session = created
            self.latest_gaze_by_stage = {}
            self.stage_gaze_offset_by_stage = {created.current_stage: 0.0}
            self.latest_raw_gaze_total_s = 0.0
            self.latest_metrics = None
            self.metric_events = []
            self.asr_events = []
            self.gaze_events = []
            self.last_asr_received_at_ms = preserved_last_asr_ms
            self.last_gaze_received_at_ms = preserved_last_gaze_ms
            self.stage_changed_at_ms = now_ms()
            self.stage_version = 1
            self.dialogue_state = {
                "formal_step": "formal_self_intro_wait",
                "star_questions": [],
                "star_index": 0,
                "self_intro_prompt": "",
            }
            policy = self._build_policy_from_condition(condition_id)
            self._question_provider = LLMQuestionProvider(self._llm, policy=policy, main_count=4)
            self._feedback_provider = LLMFeedbackProvider(self._llm, policy=policy)
            self._orchestrator = DialogueOrchestrator(self.send_robot_command, self._question_provider, self._feedback_provider)

        # 严格满足“启动 Web 后尽早触发 LLM API”：会话开始后立即做一次预热调用。
        self.ensure_llm_prewarm()

        return self.status()

    def ensure_llm_prewarm(self) -> dict:
        with self._lock:
            if self._llm_prewarm_done:
                return dict(self._llm_prewarm_status)

        status = {"ok": False, "message": "unknown", "timestamp_ms": now_ms()}
        try:
            _ = self._llm.chat_completion_text(
                system_prompt="你是模拟面试助手。仅回复‘预热成功’四个字。",
                user_prompt="请回复：预热成功",
            )
            status = {"ok": True, "message": "llm_prewarm_success", "timestamp_ms": now_ms()}
        except Exception as exc:  # noqa: BLE001
            status = {"ok": False, "message": "llm_prewarm_failed:%s" % str(exc), "timestamp_ms": now_ms()}

        with self._lock:
            self._llm_prewarm_done = True
            self._llm_prewarm_status = status
        return dict(status)

    def _condition_reactive_actions(self, stage: str) -> List[dict]:
        with self._lock:
            s = self.session
        if s is None:
            return []

        actions: List[dict] = []

        # backchanneling 维度：积极=更多肯定动作；消极=更少肯定/更审慎动作。
        if s.backchanneling_type == "positive":
            resp = self.send_robot_command("nod", {"count": 2})
            actions.append({"command": "nod", "count": 2, "ok": str(resp.get("status", "")).lower() == "ok", "tag": "cond_positive_backchannel"})
            resp2 = self.send_robot_command("gesture", {"name": "approval_nod"})
            actions.append({"command": "gesture", "name": "approval_nod", "ok": str(resp2.get("status", "")).lower() == "ok", "tag": "cond_positive_gesture"})
        else:
            resp = self.send_robot_command("nod", {"count": 1})
            actions.append({"command": "nod", "count": 1, "ok": str(resp.get("status", "")).lower() == "ok", "tag": "cond_negative_backchannel"})
            resp2 = self.send_robot_command("gesture", {"name": "disapproval_shake"})
            actions.append({"command": "gesture", "name": "disapproval_shake", "ok": str(resp2.get("status", "")).lower() == "ok", "tag": "cond_negative_gesture"})

        # persona 维度：压力型在正式阶段维持更强压迫感注视；鼓励型保持中性注视。
        if stage == "formal_interview" and s.persona_style == "pressure":
            resp3 = self.send_robot_command("gaze", {"target": "user"})
            actions.append({"command": "gaze", "target": "user", "ok": str(resp3.get("status", "")).lower() == "ok", "tag": "cond_pressure_gaze"})
        elif s.persona_style == "encouraging":
            resp3 = self.send_robot_command("gesture", {"name": "encourage_open_palm"})
            actions.append({"command": "gesture", "name": "encourage_open_palm", "ok": str(resp3.get("status", "")).lower() == "ok", "tag": "cond_encouraging_gesture"})

        return actions

    def end_session(self) -> dict:
        with self._lock:
            if self.session is None:
                raise ValueError("session_not_started")
            self.session.ended_at_ms = now_ms()
        return self.status()

    def set_stage(self, stage: str) -> dict:
        if stage not in STAGES:
            raise ValueError("invalid_stage")
        with self._lock:
            if self.session is None:
                raise ValueError("session_not_started")
            self.session.current_stage = stage
            self.session.stage_history.append({"stage": stage, "timestamp_ms": now_ms(), "source": "manual_switch"})
            self.stage_changed_at_ms = now_ms()
            self.stage_version += 1
            if stage not in self.stage_gaze_offset_by_stage:
                # Gaze 推送器可能按“全局累计时长”上报；切阶段时记下偏移，得到本阶段净时长。
                self.stage_gaze_offset_by_stage[stage] = self.latest_raw_gaze_total_s
        return self.status()

    def enter_stage(self, stage: str) -> dict:
        """切换到指定阶段，并立即下发该阶段入场动作（由 Web 完整控制流程）。"""
        self.set_stage(stage)

        script = STAGE_ENTRY_SCRIPT.get(stage, {})
        actions = []

        speak_text = str(script.get("speak", "")).strip()
        if speak_text:
            resp = self.send_robot_command("speak", {"text": speak_text})
            actions.append({"command": "speak", "ok": str(resp.get("status", "")).lower() == "ok", "response": resp})

        if bool(script.get("nod", False)):
            resp = self.send_robot_command("nod", {"count": 1})
            actions.append({"command": "nod", "ok": str(resp.get("status", "")).lower() == "ok", "response": resp})

        if stage == "formal_interview":
            # formal 子状态机：先发1分钟自我介绍提示。
            prompt = self._get_self_intro_prompt()
            self._set_formal_state(step="formal_self_intro_wait", self_intro_prompt=prompt, star_questions=[], star_index=0)
            resp = self.send_robot_command("speak", {"text": prompt})
            actions.append({"command": "speak", "ok": str(resp.get("status", "")).lower() == "ok", "response": resp, "tag": "formal_self_intro_prompt"})

        if stage == "closing_and_questionnaire":
            resp = self.send_robot_command("reset_posture", {})
            actions.append({"command": "reset_posture", "ok": str(resp.get("status", "")).lower() == "ok", "response": resp})

        # 阶段进入时按条件附加动作差异（persona × backchanneling）。
        try:
            actions.extend(self._condition_reactive_actions(stage=stage))
        except Exception as exc:  # noqa: BLE001
            actions.append({"command": "condition_actions", "ok": False, "error": str(exc)})

        s = self.status()
        s["stage_entry_actions"] = actions
        return s

    def ingest_gaze(self, payload: dict) -> dict:
        # 以 Web 控制台当前阶段为准，确保主试切换阶段后统计口径一致。
        stage = self._current_stage_or_default()
        reported_stage = str(payload.get("stage", "")).strip()
        raw_total_s = max(0.0, float(payload.get("gaze_contact_s", 0.0) or 0.0))
        ts_ms = int(payload.get("timestamp_ms", now_ms()))

        with self._lock:
            if self.stage_changed_at_ms > 0 and ts_ms < (self.stage_changed_at_ms - STAGE_SYNC_GRACE_MS):
                return {"ok": True, "dropped": True, "reason": "stale_before_stage_switch"}

            if stage not in self.stage_gaze_offset_by_stage:
                self.stage_gaze_offset_by_stage[stage] = raw_total_s

            baseline = self.stage_gaze_offset_by_stage.get(stage, 0.0)
            stage_gaze_s = max(0.0, raw_total_s - baseline)
            self.latest_raw_gaze_total_s = raw_total_s

            event = {
                "stage": stage,
                "reported_stage": reported_stage,
                "gaze_contact_s": round(stage_gaze_s, 3),
                "raw_gaze_contact_s": round(raw_total_s, 3),
                "timestamp_ms": ts_ms,
            }
            self.latest_gaze_by_stage[stage] = event
            self.gaze_events.append(event)
            self.last_gaze_received_at_ms = now_ms()
        return {"ok": True}

    def ingest_asr(self, payload: dict) -> dict:
        text = str(payload.get("text", "")).strip()
        if not text:
            raise ValueError("empty_asr_text")

        # 心跳包只用于连通性保活，不参与实验指标计算。
        is_heartbeat = bool(payload.get("heartbeat", False)) or text == "<heartbeat>"
        if is_heartbeat:
            with self._lock:
                self.last_asr_received_at_ms = now_ms()
            return {"ok": True, "heartbeat": True}

        stage = self._current_stage_or_default()
        print(f"[DEBUG] ingest_asr: stage={stage}, text={text[:50]}...")
        duration = float(payload.get("speech_duration_s", 0.0) or 0.0)
        duration = duration if duration > 0 else max(1.0, len(text) * 0.18)

        with self._lock:
            ts_ms = int(payload.get("timestamp_ms", now_ms()))
            if self.stage_changed_at_ms > 0 and ts_ms < (self.stage_changed_at_ms - STAGE_SYNC_GRACE_MS):
                return {"ok": True, "dropped": True, "reason": "stale_before_stage_switch"}

            gaze = self.latest_gaze_by_stage.get(stage, {}).get("gaze_contact_s", 0.0)
            fluency = compute_fluency_metrics(text, duration)
            metrics = {
                "timestamp_ms": ts_ms,
                "stage": stage,
                "text": text,
                "speech_duration_s": round(duration, 3),
                "gaze_contact_s": round(float(gaze), 3),
                "speech_rate_cpm": fluency["speech_rate_cpm"],
                "disfluency_ratio": fluency["disfluency_ratio"],
                "pause_ratio": fluency["pause_ratio"],
                "repetition_ratio": fluency["repetition_ratio"],
                "self_correction_ratio": fluency["self_correction_ratio"],
                "fluency_score": fluency["fluency_score"],
                "gaze_contact_ratio": compute_gaze_contact_ratio(gaze, duration),
                "source": "asr_realtime",
            }
            self.latest_metrics = metrics
            self.metric_events.append(metrics)
            self.asr_events.append({
                "timestamp_ms": metrics["timestamp_ms"],
                "stage": stage,
                "text": text,
                "speech_duration_s": round(duration, 3),
            })
            self.last_asr_received_at_ms = now_ms()

        # 在锁外触发对话动作，避免阻塞数据写入。
        print(f"[DEBUG] Calling _handle_dialogue_after_asr...")
        dialogue = self._handle_dialogue_after_asr(stage=stage, user_text=text)
        print(f"[DEBUG] Dialogue result: triggered={dialogue.get('triggered')}, reason={dialogue.get('reason')}")
        return {"ok": True, "dialogue": dialogue}

    def _set_formal_state(self, step: str, self_intro_prompt: str, star_questions: List[str], star_index: int) -> None:
        with self._lock:
            self.dialogue_state = {
                "formal_step": step,
                "self_intro_prompt": self_intro_prompt,
                "star_questions": list(star_questions),
                "star_index": int(star_index),
            }

    def _get_self_intro_prompt(self) -> str:
        provider = self._question_provider
        if provider is None:
            return "请你用约1分钟做自我介绍，重点说明经历、优势和你期待的实习方向。"
        return provider.get_self_intro_prompt()

    def _get_main_questions(self) -> List[str]:
        provider = self._question_provider
        if provider is None:
            return [
                "请用 STAR 结构介绍一个你主导推进并最终落地的项目经历。",
                "请回忆一次团队协作冲突，你当时如何沟通并推动问题解决？",
                "面对高压 deadline 时，你如何安排优先级并保证交付质量？",
                "请分享一次你快速学习新技能并应用到任务中的经历。",
            ]
        return provider.get_main_questions()

    def _feedback_text(self, answer_text: str) -> str:
        provider = self._feedback_provider
        if provider is None:
            return "收到。建议你补充更具体的行动和量化结果。"
        return provider.feedback_for_answer(answer_text)

    def _handle_dialogue_after_asr(self, stage: str, user_text: str) -> dict:
        if not hasattr(self, "_orchestrator"):
            print(f"[DEBUG] Creating DialogueOrchestrator, feedback_provider={self._feedback_provider}")
            self._orchestrator = DialogueOrchestrator(self.send_robot_command, self._question_provider, self._feedback_provider)
        try:
            with self._lock:
                local_state = dict(self.dialogue_state)
            print(f"[DEBUG] Calling orchestrator.on_asr(stage={stage})")
            result = self._orchestrator.on_asr(stage=stage, user_text=user_text, dialogue_state=local_state)
            if bool(result.get("triggered", False)):
                cond_actions = self._condition_reactive_actions(stage=stage)
                result_actions = result.get("actions", [])
                if isinstance(result_actions, list):
                    result_actions.extend(cond_actions)
                    result["actions"] = result_actions
                else:
                    result["actions"] = cond_actions
            next_state = result.get("next_state")
            if isinstance(next_state, dict):
                self._set_formal_state(
                    step=str(next_state.get("formal_step", "formal_self_intro_wait")),
                    self_intro_prompt=str(next_state.get("self_intro_prompt", "")),
                    star_questions=list(next_state.get("star_questions", [])),
                    star_index=int(next_state.get("star_index", 0)),
                )
            return result
        except Exception as exc:  # noqa: BLE001
            return {"triggered": False, "reason": "dialogue_error:%s" % str(exc)}

    def status(self) -> dict:
        with self._lock:
            session = None
            if self.session is not None:
                session = {
                    "session_id": self.session.session_id,
                    "participant_id": self.session.participant_id,
                    "participant_name": self.session.participant_name,
                    "condition_id": self.session.condition_id,
                    "persona_style": self.session.persona_style,
                    "backchanneling_type": self.session.backchanneling_type,
                    "started_at_ms": self.session.started_at_ms,
                    "ended_at_ms": self.session.ended_at_ms,
                    "current_stage": self.session.current_stage,
                    "stage_history": list(self.session.stage_history),
                }

            avg = self._aggregate_metrics_locked()
            latest_asr = self.asr_events[-1] if self.asr_events else None
            latest_gaze = self.gaze_events[-1] if self.gaze_events else None
            recent_metrics = self.metric_events[-20:]
            last_asr_ms = self.last_asr_received_at_ms
            last_gaze_ms = self.last_gaze_received_at_ms

            result = {
                "ok": True,
                "session": session,
                "conditions": CONDITION_MATRIX,
                "stages": STAGES,
                "latest_metrics": self.latest_metrics,
                "latest_asr": latest_asr,
                "latest_gaze": latest_gaze,
                "recent_metrics": recent_metrics,
                "aggregate_metrics": avg,
                "aggregate_metrics_by_stage": self._aggregate_metrics_by_stage_locked(),
                "counts": {
                    "metric_events": len(self.metric_events),
                    "asr_events": len(self.asr_events),
                    "gaze_events": len(self.gaze_events),
                },
                "dialogue_state": dict(self.dialogue_state),
                "llm_prewarm": dict(self._llm_prewarm_status),
            }

        result["connectivity"] = self._build_connectivity(last_asr_ms=last_asr_ms, last_gaze_ms=last_gaze_ms)
        return result

    def export_session(self) -> dict:
        with self._lock:
            if self.session is None:
                raise ValueError("session_not_started")
            snapshot = {
                "session": self.session,
                "metric_events": list(self.metric_events),
                "asr_events": list(self.asr_events),
                "gaze_events": list(self.gaze_events),
                "aggregate": self._aggregate_metrics_locked(),
                "aggregate_by_stage": self._aggregate_metrics_by_stage_locked(),
            }

        if not os.path.isdir(self.export_dir):
            os.makedirs(self.export_dir)

        base_name = "%s_%s" % (snapshot["session"].participant_id or "P_UNKNOWN", snapshot["session"].session_id)
        xlsx_path = os.path.join(self.export_dir, "%s.xlsx" % base_name)
        csv_path = os.path.join(self.export_dir, "%s.csv" % base_name)

        try:
            self._export_xlsx(snapshot, xlsx_path)
            return {"ok": True, "file_path": xlsx_path, "file_type": "xlsx"}
        except Exception:
            self._export_csv(snapshot, csv_path)
            return {
                "ok": True,
                "file_path": csv_path,
                "file_type": "csv",
                "note": "openpyxl_unavailable_or_xlsx_failed_fallback_to_csv",
            }

    def send_robot_command(self, command_name: str, payload: dict) -> dict:
        with self._lock:
            session = self.session

        req = {
            "protocol_version": "1.0",
            "request_id": "REQ_%s" % uuid.uuid4().hex[:12],
            "timestamp_ms": now_ms(),
            "session_id": session.session_id if session else "web_console_session",
            "participant_id": session.participant_id if session else "web_console_participant",
            "condition_id": session.condition_id if session else "web_console_condition",
            "turn_id": "WEB%03d" % (len(self.metric_events) + 1),
            "command": command_name,
            "payload": payload or {},
            "timeout_ms": 5000,
            "retry_count": 0,
        }
        body = json.dumps(req, ensure_ascii=False).encode("utf-8")
        print(f"[DEBUG] Sending robot command: {command_name}, payload={payload}")
        try:
            http_req = request.Request(
                url=self.robot_server_url,
                data=body,
                method="POST",
                headers={"Content-Type": "application/json; charset=utf-8"},
            )
            with request.urlopen(http_req, timeout=6.0) as resp:
                raw = resp.read().decode("utf-8")
            result = json.loads(raw)
            print(f"[DEBUG] Robot response: status={result.get('status')}, message={result.get('message')}")
            return result
        except Exception as e:
            print(f"[DEBUG] Robot command failed: {e}")
            return {"status": "error", "message": str(e)}

    def _build_connectivity(self, last_asr_ms: int, last_gaze_ms: int) -> dict:
        now = now_ms()
        asr_age = None if last_asr_ms <= 0 else max(0, now - last_asr_ms)
        gaze_age = None if last_gaze_ms <= 0 else max(0, now - last_gaze_ms)

        asr_ok = (asr_age is not None) and (asr_age <= ASR_STALE_TIMEOUT_MS)
        gaze_ok = (gaze_age is not None) and (gaze_age <= GAZE_STALE_TIMEOUT_MS)
        robot_health = self._get_robot_server_health()

        warnings = []
        if not asr_ok:
            warnings.append("ASR 推送未连接或已超时，请检查 asr_realtime_pusher.py 与 /asr URL")
        if not gaze_ok:
            warnings.append("Gaze 推送未连接或已超时，请检查 gaze_realtime_pusher.py 与 /gaze URL")
        if not robot_health.get("ok", False):
            warnings.append("机器人 command_server 不可达，请检查 robot_server_py2/command_server.py")

        return {
            "asr": {
                "ok": asr_ok,
                "last_received_at_ms": last_asr_ms or None,
                "age_ms": asr_age,
                "stale_timeout_ms": ASR_STALE_TIMEOUT_MS,
            },
            "gaze": {
                "ok": gaze_ok,
                "last_received_at_ms": last_gaze_ms or None,
                "age_ms": gaze_age,
                "stale_timeout_ms": GAZE_STALE_TIMEOUT_MS,
            },
            "command_server": robot_health,
            "warnings": warnings,
        }

    def _get_robot_server_health(self) -> dict:
        now = now_ms()
        with self._lock:
            cached = dict(self._robot_health_cache)

        checked_at = int(cached.get("checked_at_ms", 0) or 0)
        if (now - checked_at) <= ROBOT_HEALTH_CACHE_MS:
            return cached

        fresh = self._probe_robot_server()
        with self._lock:
            self._robot_health_cache = fresh
        return fresh

    def _probe_robot_server(self) -> dict:
        req = {
            "protocol_version": "1.0",
            "request_id": "HEALTH_%s" % uuid.uuid4().hex[:10],
            "timestamp_ms": now_ms(),
            "session_id": "web_console_health",
            "participant_id": "web_console_health",
            "condition_id": "web_console_health",
            "turn_id": "HEALTH",
            "command": "ping",
            "payload": {},
            "timeout_ms": 1200,
            "retry_count": 0,
        }

        body = json.dumps(req, ensure_ascii=False).encode("utf-8")
        http_req = request.Request(
            url=self.robot_server_url,
            data=body,
            method="POST",
            headers={"Content-Type": "application/json; charset=utf-8"},
        )

        checked_at = now_ms()
        try:
            with request.urlopen(http_req, timeout=1.5) as resp:
                raw = resp.read().decode("utf-8")
            data = json.loads(raw)
            status = str(data.get("status", "")).lower()
            ok = status == "ok"
            message = "reachable" if ok else (data.get("message") or "command_server_response_error")
            return {
                "checked_at_ms": checked_at,
                "ok": ok,
                "message": str(message),
                "url": self.robot_server_url,
            }
        except (HTTPError, URLError) as exc:
            return {
                "checked_at_ms": checked_at,
                "ok": False,
                "message": "network_error:%s" % str(exc),
                "url": self.robot_server_url,
            }
        except Exception as exc:  # noqa: BLE001
            return {
                "checked_at_ms": checked_at,
                "ok": False,
                "message": "probe_failed:%s" % str(exc),
                "url": self.robot_server_url,
            }

    def _current_stage_or_default(self) -> str:
        with self._lock:
            if self.session is not None:
                return self.session.current_stage
        return "warmup"

    def _aggregate_metrics_locked(self) -> dict:
        if not self.metric_events:
            return {
                "speech_rate_cpm": 0.0,
                "disfluency_ratio": 0.0,
                "pause_ratio": 0.0,
                "repetition_ratio": 0.0,
                "self_correction_ratio": 0.0,
                "fluency_score": 0.0,
                "gaze_contact_ratio": 0.0,
            }
        n = float(len(self.metric_events))
        return {
            "speech_rate_cpm": round(sum(m["speech_rate_cpm"] for m in self.metric_events) / n, 3),
            "disfluency_ratio": round(sum(m["disfluency_ratio"] for m in self.metric_events) / n, 6),
            "pause_ratio": round(sum(float(m.get("pause_ratio", 0.0)) for m in self.metric_events) / n, 6),
            "repetition_ratio": round(sum(float(m.get("repetition_ratio", 0.0)) for m in self.metric_events) / n, 6),
            "self_correction_ratio": round(sum(float(m.get("self_correction_ratio", 0.0)) for m in self.metric_events) / n, 6),
            "fluency_score": round(sum(float(m.get("fluency_score", 0.0)) for m in self.metric_events) / n, 6),
            "gaze_contact_ratio": round(sum(m["gaze_contact_ratio"] for m in self.metric_events) / n, 6),
        }

    def _aggregate_metrics_by_stage_locked(self) -> dict:
        grouped: Dict[str, List[dict]] = {}
        for m in self.metric_events:
            grouped.setdefault(str(m.get("stage", "unknown")), []).append(m)

        out = {}
        for stage, rows in grouped.items():
            n = float(len(rows))
            out[stage] = {
                "count": int(n),
                "speech_rate_cpm": round(sum(r["speech_rate_cpm"] for r in rows) / n, 3),
                "disfluency_ratio": round(sum(r["disfluency_ratio"] for r in rows) / n, 6),
                "pause_ratio": round(sum(float(r.get("pause_ratio", 0.0)) for r in rows) / n, 6),
                "repetition_ratio": round(sum(float(r.get("repetition_ratio", 0.0)) for r in rows) / n, 6),
                "self_correction_ratio": round(sum(float(r.get("self_correction_ratio", 0.0)) for r in rows) / n, 6),
                "fluency_score": round(sum(float(r.get("fluency_score", 0.0)) for r in rows) / n, 6),
                "gaze_contact_ratio": round(sum(r["gaze_contact_ratio"] for r in rows) / n, 6),
            }
        return out

    @staticmethod
    def _export_csv(snapshot: dict, out_path: str) -> None:
        import csv

        session = snapshot["session"]
        rows = snapshot["metric_events"]
        with open(out_path, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f)
            writer.writerow(["session_id", session.session_id])
            writer.writerow(["participant_id", session.participant_id])
            writer.writerow(["participant_name", session.participant_name])
            writer.writerow(["condition_id", session.condition_id])
            writer.writerow(["persona_style", session.persona_style])
            writer.writerow(["backchanneling_type", session.backchanneling_type])
            writer.writerow([])
            writer.writerow(["stage", "count", "avg_speech_rate_cpm", "avg_disfluency_ratio", "avg_pause_ratio", "avg_repetition_ratio", "avg_self_correction_ratio", "avg_fluency_score", "avg_gaze_contact_ratio"])
            stage_agg = snapshot.get("aggregate_by_stage", {})
            for stage in sorted(stage_agg.keys()):
                row = stage_agg[stage]
                writer.writerow([stage, row.get("count", 0), row.get("speech_rate_cpm", 0.0), row.get("disfluency_ratio", 0.0), row.get("pause_ratio", 0.0), row.get("repetition_ratio", 0.0), row.get("self_correction_ratio", 0.0), row.get("fluency_score", 0.0), row.get("gaze_contact_ratio", 0.0)])

            writer.writerow([])
            writer.writerow(["timestamp_ms", "stage", "text", "speech_duration_s", "gaze_contact_s", "speech_rate_cpm", "disfluency_ratio", "pause_ratio", "repetition_ratio", "self_correction_ratio", "fluency_score", "gaze_contact_ratio"])
            for r in rows:
                writer.writerow([
                    r.get("timestamp_ms", 0),
                    r.get("stage", ""),
                    r.get("text", ""),
                    r.get("speech_duration_s", 0.0),
                    r.get("gaze_contact_s", 0.0),
                    r.get("speech_rate_cpm", 0.0),
                    r.get("disfluency_ratio", 0.0),
                    r.get("pause_ratio", 0.0),
                    r.get("repetition_ratio", 0.0),
                    r.get("self_correction_ratio", 0.0),
                    r.get("fluency_score", 0.0),
                    r.get("gaze_contact_ratio", 0.0),
                ])

    @staticmethod
    def _export_xlsx(snapshot: dict, out_path: str) -> None:
        from openpyxl import Workbook

        session = snapshot["session"]
        wb = Workbook()
        ws_summary = wb.active
        ws_summary.title = "summary"
        ws_summary.append(["field", "value"])
        ws_summary.append(["session_id", session.session_id])
        ws_summary.append(["participant_id", session.participant_id])
        ws_summary.append(["participant_name", session.participant_name])
        ws_summary.append(["condition_id", session.condition_id])
        ws_summary.append(["persona_style", session.persona_style])
        ws_summary.append(["backchanneling_type", session.backchanneling_type])
        ws_summary.append(["started_at_ms", session.started_at_ms])
        ws_summary.append(["ended_at_ms", session.ended_at_ms])
        ws_summary.append(["avg_speech_rate_cpm", snapshot["aggregate"]["speech_rate_cpm"]])
        ws_summary.append(["avg_disfluency_ratio", snapshot["aggregate"]["disfluency_ratio"]])
        ws_summary.append(["avg_pause_ratio", snapshot["aggregate"].get("pause_ratio", 0.0)])
        ws_summary.append(["avg_repetition_ratio", snapshot["aggregate"].get("repetition_ratio", 0.0)])
        ws_summary.append(["avg_self_correction_ratio", snapshot["aggregate"].get("self_correction_ratio", 0.0)])
        ws_summary.append(["avg_fluency_score", snapshot["aggregate"].get("fluency_score", 0.0)])
        ws_summary.append(["avg_gaze_contact_ratio", snapshot["aggregate"]["gaze_contact_ratio"]])

        ws_stage_agg = wb.create_sheet("stage_metrics")
        ws_stage_agg.append(["stage", "count", "avg_speech_rate_cpm", "avg_disfluency_ratio", "avg_pause_ratio", "avg_repetition_ratio", "avg_self_correction_ratio", "avg_fluency_score", "avg_gaze_contact_ratio"])
        for stage in sorted(snapshot.get("aggregate_by_stage", {}).keys()):
            row = snapshot["aggregate_by_stage"][stage]
            ws_stage_agg.append([
                stage,
                row.get("count", 0),
                row.get("speech_rate_cpm", 0.0),
                row.get("disfluency_ratio", 0.0),
                row.get("pause_ratio", 0.0),
                row.get("repetition_ratio", 0.0),
                row.get("self_correction_ratio", 0.0),
                row.get("fluency_score", 0.0),
                row.get("gaze_contact_ratio", 0.0),
            ])

        ws_metrics = wb.create_sheet("metrics")
        ws_metrics.append(["timestamp_ms", "stage", "text", "speech_duration_s", "gaze_contact_s", "speech_rate_cpm", "disfluency_ratio", "pause_ratio", "repetition_ratio", "self_correction_ratio", "fluency_score", "gaze_contact_ratio"])
        for row in snapshot["metric_events"]:
            ws_metrics.append([
                row.get("timestamp_ms", 0),
                row.get("stage", ""),
                row.get("text", ""),
                row.get("speech_duration_s", 0.0),
                row.get("gaze_contact_s", 0.0),
                row.get("speech_rate_cpm", 0.0),
                row.get("disfluency_ratio", 0.0),
                row.get("pause_ratio", 0.0),
                row.get("repetition_ratio", 0.0),
                row.get("self_correction_ratio", 0.0),
                row.get("fluency_score", 0.0),
                row.get("gaze_contact_ratio", 0.0),
            ])

        ws_stage = wb.create_sheet("stage_history")
        ws_stage.append(["stage", "timestamp_ms", "source"])
        for h in session.stage_history:
            ws_stage.append([h.get("stage", ""), h.get("timestamp_ms", 0), h.get("source", "")])

        wb.save(out_path)


class WebConsoleServer(object):
    def __init__(
        self,
        host: str = "127.0.0.1",
        port: int = 8780,
        static_dir: Optional[str] = None,
        export_dir: str = "client_py3/exports",
        robot_server_url: str = "http://127.0.0.1:8000/command",
    ):
        self.host = host
        self.port = int(port)
        self.static_dir = static_dir or os.path.join(os.path.dirname(__file__), "web_console")
        self.state = ExperimentState(export_dir=export_dir, robot_server_url=robot_server_url)
        self._server: Optional[ThreadingHTTPServer] = None

    @property
    def web_url(self) -> str:
        return "http://%s:%d" % (self.host, self.port)

    def start(self) -> None:
        self._server = ThreadingHTTPServer((self.host, self.port), self._make_handler())
        print("[INFO] web_console_started: http://%s:%d" % (self.host, self.port))
        print("[INFO] robot_server_url=%s" % self.state.robot_server_url)
        self._server.serve_forever()

    def _make_handler(self):
        state = self.state
        static_dir = self.static_dir

        class _Handler(BaseHTTPRequestHandler):
            def _send_json(self, status: int, body: dict) -> None:
                raw = json.dumps(body, ensure_ascii=False).encode("utf-8")
                self.send_response(status)
                self.send_header("Content-Type", "application/json; charset=utf-8")
                self.send_header("Content-Length", str(len(raw)))
                self.end_headers()
                self.wfile.write(raw)

            def _read_json(self) -> dict:
                size = int(self.headers.get("Content-Length", "0") or 0)
                raw = self.rfile.read(size) if size > 0 else b"{}"
                return json.loads(raw.decode("utf-8"))

            def _serve_index(self) -> None:
                # 打开 Web 首页时立刻触发一次 LLM 预热（异步，不阻塞页面加载）。
                threading.Thread(target=state.ensure_llm_prewarm, daemon=True).start()
                path = os.path.join(static_dir, "index.html")
                if not os.path.isfile(path):
                    self._send_json(404, {"ok": False, "error": "index_not_found"})
                    return
                with open(path, "rb") as f:
                    raw = f.read()
                self.send_response(200)
                self.send_header("Content-Type", "text/html; charset=utf-8")
                self.send_header("Content-Length", str(len(raw)))
                self.end_headers()
                self.wfile.write(raw)

            def do_GET(self):  # noqa: N802
                if self.path in ["/", "/index.html"]:
                    self._serve_index()
                    return
                if self.path == "/api/health":
                    self._send_json(200, {"ok": True, "service": "web_console"})
                    return
                if self.path == "/api/status":
                    self._send_json(200, state.status())
                    return
                self._send_json(404, {"ok": False, "error": "not_found"})

            def do_POST(self):  # noqa: N802
                try:
                    if self.path == "/asr":
                        payload = self._read_json()
                        self._send_json(200, state.ingest_asr(payload))
                        return

                    if self.path == "/gaze":
                        payload = self._read_json()
                        self._send_json(200, state.ingest_gaze(payload))
                        return

                    if self.path == "/api/session/start":
                        p = self._read_json()
                        result = state.start_session(
                            participant_id=str(p.get("participant_id", "")).strip(),
                            participant_name=str(p.get("participant_name", "")).strip(),
                            condition_id=str(p.get("condition_id", "")).strip(),
                        )
                        self._send_json(200, result)
                        return

                    if self.path == "/api/session/end":
                        result = state.end_session()
                        export_result = state.export_session()
                        result["export"] = export_result
                        self._send_json(200, result)
                        return

                    if self.path == "/api/session/export":
                        self._send_json(200, state.export_session())
                        return

                    if self.path == "/api/stage":
                        p = self._read_json()
                        stage = str(p.get("stage", "")).strip()
                        enter_now = bool(p.get("enter_now", True))
                        result = state.enter_stage(stage) if enter_now else state.set_stage(stage)
                        self._send_json(200, result)
                        return

                    if self.path == "/api/robot/command":
                        p = self._read_json()
                        cmd = str(p.get("command", "")).strip()
                        if not cmd:
                            self._send_json(400, {"ok": False, "error": "missing_command"})
                            return
                        payload = p.get("payload", {})
                        if payload is None:
                            payload = {}
                        if not isinstance(payload, dict):
                            self._send_json(400, {"ok": False, "error": "invalid_payload"})
                            return
                        data = state.send_robot_command(cmd, payload)
                        self._send_json(200, {"ok": True, "robot_response": data})
                        return

                    self._send_json(404, {"ok": False, "error": "not_found"})
                except ValueError as exc:
                    self._send_json(400, {"ok": False, "error": str(exc)})
                except Exception as exc:  # noqa: BLE001
                    self._send_json(500, {"ok": False, "error": "internal_error", "reason": str(exc)})

            def log_message(self, fmt: str, *args):
                del fmt, args
                return

        return _Handler


def _parse_args():
    parser = argparse.ArgumentParser(description="NAO 实验 Web 控制台")
    parser.add_argument("--host", type=str, default="127.0.0.1", help="Web 服务监听地址")
    parser.add_argument("--port", type=int, default=8780, help="Web 服务监听端口")
    parser.add_argument(
        "--robot-server-url",
        type=str,
        default="http://127.0.0.1:8000/command",
        help="Python2 command_server 的 URL",
    )
    parser.add_argument(
        "--export-dir",
        type=str,
        default=os.path.join(os.path.dirname(__file__), "exports"),
        help="实验数据导出目录（xlsx/csv）",
    )
    parser.add_argument(
        "--static-dir",
        type=str,
        default=os.path.join(os.path.dirname(__file__), "web_console"),
        help="前端静态文件目录",
    )
    return parser.parse_args()


def main():
    args = _parse_args()
    server = WebConsoleServer(
        host=args.host,
        port=args.port,
        static_dir=args.static_dir,
        export_dir=args.export_dir,
        robot_server_url=args.robot_server_url,
    )
    server.start()


if __name__ == "__main__":
    main()
